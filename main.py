import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import numpy as np
import re
import openpyxl

from db import DataBase, CFG
import classes
import storage


class App:
    """Main window"""
    tk_window = None

    def __init__(self):
        self.tk_window = tk.Tk()
        # Configure the root window
        self.tk_window.title('Sand granulometric composition')
        self.tk_window.state('zoomed')
        test_img = tk.PhotoImage(file='icon.png')
        self.tk_window.iconphoto(True, test_img)
        # Create notebook
        self.notebook = ttk.Notebook(self.tk_window)
        self.notebook.pack(fill='both', expand=True)
        # Create tabs
        self.__sample = Sample(self.notebook)
        # self.__compare_samples = CompareSamples(self.notebook)
        self.__compare_samples = classes.Table(self.notebook, DataBase(), columns=DataBase().headers)
        self.__settings = Settings(self.notebook)
        # Add tabs to notebook

        self.notebook.add(self.__sample.frame, text='Add the sample')
        self.notebook.add(self.__compare_samples.frame, text='Compare samples')
        self.notebook.add(self.__settings.frame, text='Settings')


class SampleInfoBlock:
    """Block for gathering information about sample in Sample tab"""
    # Create input block
    def __init__(self, root, add_btn):
        self.__add_btn = add_btn
        self.__info_frame = tk.Frame(root)
        self.__info_frame.pack(anchor='nw')
        # Collector
        self.__collector_label = tk.Label(self.__info_frame, text='Collector')
        self.__collector_label.grid(row=0, column=0, sticky='w')
        self.__collector_combobox = ttk.Combobox(self.__info_frame, width=37)
        self.__collector_combobox.grid(row=0, column=1)
        # Performer
        self.__performer_label = tk.Label(self.__info_frame, text='Performer')
        self.__performer_label.grid(row=1, column=0, sticky='w')
        self.__performer_combobox = ttk.Combobox(self.__info_frame, width=37)
        self.__performer_combobox.grid(row=1, column=1)
        # Sample
        self.__sample_label = tk.Label(self.__info_frame, text='Sample name')
        self.__sample_label.grid(row=2, column=0, sticky='w')
        self.__s_vcmd = (self.__info_frame.register(self.__validate_sample), '%P')
        self.__sample_entry = tk.Entry(self.__info_frame, validate='focusout', validatecommand=self.__s_vcmd, width=40)
        self.__sample_entry.grid(row=2, column=1)
        # Location
        self.__location_label = tk.Label(self.__info_frame, text='Location name')
        self.__location_label.grid(row=3, column=0, sticky='w')
        self.__location_combobox = ttk.Combobox(self.__info_frame, width=37)
        self.__location_combobox.grid(row=3, column=1)
        # Zone
        self.__zone_label = tk.Label(self.__info_frame, text='Zone')
        self.__zone_label.grid(row=4, column=0, sticky='w')
        self.__zone_combobox = ttk.Combobox(self.__info_frame, width=37)
        self.__zone_combobox.grid(row=4, column=1, )
        # Latitude
        self.__lat_label = tk.Label(self.__info_frame, text='Latitude')
        self.__lat_label.grid(row=5, column=0, sticky='w')
        self.__lat_entry = tk.Entry(self.__info_frame, width=40)
        self.__lat_entry.grid(row=5, column=1)
        # Longitude
        self.__lon_label = tk.Label(self.__info_frame, text='Longitude')
        self.__lon_label.grid(row=6, column=0, sticky='w')
        self.__lon_entry = tk.Entry(self.__info_frame, width=40)
        self.__lon_entry.grid(row=6, column=1)
        # Sampling date
        self.__date_label = tk.Label(self.__info_frame, text='Sampling date')
        self.__date_label.grid(row=7, column=0, sticky='w')
        self.__d_vcmd = (self.__info_frame.register(self.__vali_date), '%P')
        self.__date_entry = tk.Entry(self.__info_frame, width=40, validate='focusout', validatecommand=self.__d_vcmd)
        self.__date_entry.insert(tk.END, 'yyyy-mm-dd')
        self.__date_entry.grid(row=7, column=1)
        self.upd_combobox_values()

    @property
    def sample(self):
        return self.__sample_entry.get()

    def upd_combobox_values(self):
        persons = DataBase().update('person')
        self.__collector_combobox['values'] = persons
        self.__performer_combobox['values'] = persons
        self.__zone_combobox['values'] = DataBase().update('zone')
        self.__location_combobox['values'] = DataBase().update('location')

    # def __upd_persons(self) -> None:
    #     """ Update collector and performer combobox values from database"""
    #     persons = DataBase().update('person')
    #     self.__collector_combobox['values'] = persons
    #     self.__performer_combobox['values'] = persons
    #
    # def __upd_zones(self) -> None:
    #     """Update zone combobox values from database"""
    #     self.__zone_combobox['values'] = DataBase().update('zone')
    #
    # def __upd_locations(self) -> None:
    #     """Update location combobox values from database"""
    #     self.__location_combobox['values'] = DataBase().update('location')

    def gather_info(self) -> storage.SampleData:
        """:return: SampleData dataclass with sample information"""
        return storage.SampleData(
            collector=self.__collector_combobox.get(),
            performer=self.__performer_combobox.get(),
            sample=self.__sample_entry.get(),
            location=self.__location_combobox.get(),
            zone=self.__zone_combobox.get(),
            lat=self.__lat_entry.get(),
            lon=self.__lon_entry.get(),
            sampling_date=self.__date_entry.get()
        )

    def __vali_date(self, value: str) -> bool:
        """
        Validation of date

        :param value: date_entry string
        :return: True if value is yyyy.mm.dd date, else False
        """
        pattern = r'[1|2][0-9]{3}-(0[1-9]|1[0-2])-(0[1-9]|[1-2][0-9]|3[0-1])'
        if re.fullmatch(pattern, value) is None:
            self.__date_entry.config(fg='red')
            self.__add_btn.config(state='disabled', text='Please use yyyy-mm-dd date format.')
            return False
        else:
            self.__date_entry.config(fg='black')
            self.__add_btn.config(state='normal', text='Add')
            return True

    def __validate_sample(self, value: str) -> bool:
        """
        Validation of sample

        :param value: sample_entry string
        :return: False if sample name already exists in database
        """
        if value in DataBase().update('sample'):
            self.__sample_entry.config(fg='red')
            self.__add_btn.config(state='disabled', text=f'Sample "{value}" already exists')
            return False
        else:
            self.__sample_entry.config(fg='black')
            self.__add_btn.config(state='normal', text='Add')
            return True


class WeightsBlock:
    """Block for gathering weights and calculation of indices"""
    def __init__(self,  root):
        # Fractions and weights: Label and table of entry widgets
        self.__fractions_label = tk.Label(root, text='Fractions and weights:')
        self.__fractions_label.pack()
        self.__weight_entry = classes.MultipleEntry(root, Fractions().schemes[CFG().def_fract])
        self.__weight_entry.frame.pack()
        # Update entry table when tab is focused
        root.bind('<FocusIn>', self.__upd_fw)

    def __upd_fw(self, event) -> None:
        """Update fractions entry if scheme was modified in Settings"""
        if len(self.__weight_entry) != len(Fractions().schemes[CFG().def_fract]):
            self.__weight_entry.upd(Fractions().schemes[CFG().def_fract])

    def compute(self):
        """:return: fractions and cumulative weights, indices values in IndicesData dataclass"""
        fractions = Fractions().schemes[CFG().def_fract]
        weights = np.array(self.__weight_entry.get(), dtype=float)
        cumulative_weights, indices = self.calculate_indices(fractions, weights)
        return fractions, cumulative_weights, indices

    @staticmethod
    def calculate_indices(fractions: np.array, weights: np.array):
        """
        Calculations for compute and import methods

        :return: Cumulative weights array and IndicesData dataclass with indices
        """
        cumulative_weights = np.cumsum(100 * weights / sum(weights)).round(CFG().rnd_frac)
        percentiles = (5, 16, 25, 50, 68, 75, 84, 95)
        phi = dict(zip(percentiles, np.interp(percentiles, cumulative_weights, fractions)))
        return cumulative_weights, storage.IndicesData(
            MdPhi=round(phi[50], CFG().rnd_ind),
            Mz=round((phi[16] + phi[50] + phi[84]) / 3, CFG().rnd_ind),
            QDPhi=round((phi[75] - phi[25]) / 2, CFG().rnd_ind),
            sigma_1=round((phi[84] - phi[16]) / 4 + (phi[95] - phi[5]) / 6.6, CFG().rnd_ind),
            SkqPhi=round((phi[25] + phi[75] - phi[50]) / 2, CFG().rnd_ind),
            Sk_1=round((phi[16] + phi[84] - 2 * phi[50]) / (2 * (phi[84] - phi[16])) +
                       (phi[5] + phi[95] - 2 * phi[50]) / (2 * (phi[95] - phi[5])), CFG().rnd_ind),
            KG=round((phi[95] - phi[5]) / (2.44 * (phi[75] - phi[25])), CFG().rnd_ind),
            SD=round(phi[68], CFG().rnd_ind)
        )


class Sample:
    """Sample tab"""
    frame = None

    def __init__(self, container):
        self.frame = ttk.Frame(container)
        # Create left frame
        self.__left_frame = tk.Frame(self.frame)
        self.__left_frame.pack(side=tk.LEFT, anchor='nw', fill='y')
        # Frame for buttons
        self.__btn_frame = tk.Frame(self.__left_frame)
        # "Check" button
        self.__upd_btn = tk.Button(self.__btn_frame, text='Check', command=self.__check_sample)
        self.__upd_btn.grid(row=0, column=0)
        # "Add" button
        self.__add_btn = tk.Button(self.__btn_frame, text='Add', command=self.__add_btn_cmd)
        self.__add_btn.grid(row=0, column=1)
        self.__info = SampleInfoBlock(self.__left_frame, self.__add_btn)
        self.__weights_block = WeightsBlock(self.__left_frame)
        self.__btn_frame.pack()
        # Create indices table
        self.__indices_table = ttk.Treeview(self.__left_frame, columns=tuple(range(8)), show='headings', height=1)
        self.__indices_table['columns'] = list(range(8))
        for i in range(8):
            self.__indices_table.column(i, width=50)
            self.__indices_table.heading(i, text=DataBase().headers[9 + i])
        self.__indices_table.pack(pady=20)
        # Create fractions table
        self.__fractions_table = ttk.Treeview(self.__left_frame, columns=('0', '1'), show='headings', height=10)
        self.__fractions_table.pack(anchor='n')
        self.__fractions_table.heading('0', text='Fractions')
        self.__fractions_table.heading('1', text='Weights')
        # Create right frame
        self.__right_frame = tk.Frame(self.frame)
        self.__right_frame.pack(anchor='nw')
        # Create cumulative curve plot in right frame
        self.__curve = classes.Curve(self.__right_frame)
        self.__curve.pack(padx=20)
        self.__import_button = tk.Button(self.__left_frame, text='Import Excel file', command=self.__import_excel)
        self.__import_button.pack()

    # TODO: rename methods according to functionality
    # TODO: Refactoring according to SOLID principals

    def __check_sample(self) -> None:
        """
        Insert calculated data in tables and plot

        :return: None
        """
        fractions, cumulative_weights, ind = self.__weights_block.compute()
        # Update indices table
        self.__indices_table.delete(*self.__indices_table.get_children())
        self.__indices_table.insert('', 'end', values=ind.get())
        # Update fractions and cumulative weights table
        self.__fractions_table.delete(*self.__fractions_table.get_children())
        for i in range(len(fractions)):
            self.__fractions_table.insert('', 'end', values=(fractions[i], cumulative_weights[i]))
        # Update plot
        self.__curve.upd([fractions], [cumulative_weights], [self.__info.sample])

    def __add_btn_cmd(self) -> None:
        """Add info and calculated parameters into database"""
        DataBase().add(*self.__weights_block.compute(), self.__info.gather_info())
        self.__info.upd_combobox_values()

    def __import_excel(self) -> None:
        """Import Excel workbook to database"""
        print('Import Excel file...')
        # Open Excel workbook
        wb = openpyxl.load_workbook(filedialog.askopenfilename(title='Open Excel book',
                                                               filetypes=[("Excel files", ".xlsx .xls")]))
        # Open active sheet
        worksheet = wb.active
        fractions = []
        # Read fractions
        for col in worksheet.iter_cols(9, worksheet.max_column):
            fractions.append(col[0].value)
        fractions = np.array(fractions, dtype=float)
        # Read rows
        for i in range(1, worksheet.max_row):
            row = []
            # Read each column in row, write to row
            for col in worksheet.iter_cols(1, worksheet.max_column):
                row.append(col[i].value)
            if self.__info.sample not in DataBase().update('sample'):
                if messagebox.askyesno(title='Conflict',
                                       message=f'Sample{row[2]} already exists in database. Do you want to rename it'
                                               f' to {row[2]}-{worksheet.title}?'):
                    row[2] = f'{row[2]}-{worksheet.title}'
            # Write weights
            weights = np.array(row[8:worksheet.max_column], dtype=float)
            # Write sample information to SampleData dataclass
            info = storage.SampleData(*row[0:7], row[7].strftime("%Y-%m-%d"))
            # Calculate cumulative weights and indices
            cumulative_weights, indices = self.__weights_block.calculate_indices(fractions, weights)
            DataBase().add(fractions, cumulative_weights, indices, info)


class Settings:
    """
    Settings tab
    """
    frame = None

    def __init__(self, container):
        self.frame = ttk.Frame(container)
        # Little frame to align blocks
        self.__main = tk.Frame(self.frame)
        self.__main.pack()
        # General settings LabelFrame
        self.__set_lb = tk.LabelFrame(self.__main, text='General settings')
        self.__set_lb.pack(fill='both')
        # Indices rounding
        self.__ind_rnd_label = tk.Label(self.__set_lb, text='Indices rounding')
        self.__ind_rnd_label.grid(row=0, column=0)
        self.__ind_rnd = ttk.Combobox(self.__set_lb, values=('1', '2', '3', '4'), width=38)
        self.__ind_rnd.insert(tk.END, str(CFG().rnd_ind))
        self.__ind_rnd.grid(row=0, column=1)
        # Weights rounding
        self.__frac_rnd_label = tk.Label(self.__set_lb, text='Weights rounding')
        self.__frac_rnd_label.grid(row=1, column=0)
        self.__frac_rnd = ttk.Combobox(self.__set_lb, values=('1', '2', '3', '4'), width=38)
        self.__frac_rnd.insert(tk.END, str(CFG().rnd_frac))
        self.__frac_rnd.grid(row=1, column=1)
        # Separator. Currently, used only in next two blocks.
        self.__sep_label = tk.Label(self.__set_lb, text='Separator')
        self.__sep_label.grid(row=3, column=0)
        self.__sep_entry = ttk.Combobox(self.__set_lb, values=(' ', ', '), width=38)
        self.__sep_entry.insert(tk.END, CFG().sep)
        self.__sep_entry.grid(row=3, column=1)
        # Fractions scheme selection
        self.__fract_label = tk.Label(self.__set_lb, text='Select fractions scheme')
        self.__fract_label.grid(row=4, column=0)
        self.__fract_cb = ttk.Combobox(self.__set_lb, postcommand=self.__upd_fractions, width=38)
        self.__fract_cb.insert(tk.END, CFG().def_fract)
        self.__fract_cb.grid(row=4, column=1)
        # Apply Button
        self.__apply_button = tk.Button(self.__set_lb, text='Apply settings', command=self.__apply)
        self.__apply_button.grid(row=5, column=0)
        # Converter
        self.__conv = tk.LabelFrame(self.__main, text='Converter')
        self.__conv.pack(fill='both')
        # Input and output
        self.__inp_label = tk.Label(self.__conv, text='Input')
        self.__inp_label.grid(row=0, column=0)
        self.__inp = tk.Entry(self.__conv, width=48)
        self.__inp.grid(row=0, column=1)
        self.__outp_label = tk.Label(self.__conv, text='Output')
        self.__outp_label.grid(row=1, column=0)
        self.__outp = tk.Entry(self.__conv, width=48)
        self.__outp.grid(row=1, column=1)
        # Convert mode
        self.__convert_options = ['D(mm) to D(φ)', 'D(mm) to D(μ)', 'D(φ) to D(mm)', 'D(φ) to D(μ)', 'D(μ) to D(mm)',
                                  'D(μ) to D(φ)']
        self.__convert_combobox = ttk.Combobox(self.__conv, values=self.__convert_options)
        self.__convert_combobox.grid(row=3, column=1)
        # "Convert" button
        self.__convert_button = tk.Button(self.__conv, text='Convert', command=self.__convert)
        self.__convert_button.grid(row=3, column=0)
        # Add fractions
        self.__fract_lf = tk.LabelFrame(self.__main, text='Add fractions scheme')
        self.__fract_lf.pack(fill='both')
        # Name
        self.__fract_name_label = tk.Label(self.__fract_lf, text='Name:')
        self.__fract_name_label.grid(row=0, column=0)
        self.__fract_name = tk.Entry(self.__fract_lf, width=49)
        self.__fract_name.grid(row=0, column=1)
        # Scheme
        self.__fract_sch_label = tk.Label(self.__fract_lf, text='Scheme:')
        self.__fract_sch_label.grid(row=1, column=0)
        self.__fract_sch = tk.Entry(self.__fract_lf, width=49)
        self.__fract_sch.grid(row=1, column=1)
        # Add button
        self.__fract_add_button = tk.Button(self.__fract_lf, text='Add', command=self.__add_fractions)
        self.__fract_add_button.grid(row=2, column=0)

    def __convert(self) -> None:
        """Converter function. Insert into output converted values."""
        inp = np.array(self.__inp.get().split(CFG().sep), dtype=float)
        if self.__convert_combobox.get() == 'D(mm) to D(φ)':  # -log2(x, mm)
            self.__outp.delete(0, tk.END)
            self.__outp.insert(tk.END, f'{CFG().sep}'.join(map(str, np.around(-np.log2(inp), 2))))
        if self.__convert_combobox.get() == 'D(mm) to D(μ)':  # x,mm * 1000 / 2
            self.__outp.delete(0, tk.END)
            self.__outp.insert(tk.END, f'{CFG().sep}'.join(map(str, inp * 1000)))
        if self.__convert_combobox.get() == 'D(φ) to D(mm)':  # 1 / (2^x,phi)
            self.__outp.delete(0, tk.END)
            self.__outp.insert(tk.END, f'{CFG().sep}'.join(map(str, np.around(1 / (2 ** inp), 2))))
        if self.__convert_combobox.get() == 'D(φ) to D(μ)':  # 1000 / 2^x,phi
            self.__outp.delete(0, tk.END)
            self.__outp.insert(tk.END, f'{CFG().sep}'.join(map(str, np.around(1000 / (2 ** inp), 2))))
        if self.__convert_combobox.get() == 'D(μ) to D(φ)':  # -log2(x,mu / 1000)
            self.__outp.delete(0, tk.END)
            self.__outp.insert(tk.END, f'{CFG().sep}'.join(map(str, np.around(-np.log2(0.001 * inp), 2))))
        if self.__convert_combobox.get() == 'D(μ) to D(mm)':  # x,mu / 1000
            self.__outp.delete(0, tk.END)
            self.__outp.insert(tk.END, f'{CFG().sep}'.join(map(str, np.around(0.001 * inp, 2))))

    def __apply(self) -> None:
        """Apply settings: write it to the config.txt, update cfg"""
        CFG().apply_settings(self.__sep_entry.get(),
                             self.__ind_rnd.get(),
                             self.__frac_rnd.get(),
                             self.__fract_cb.get()
                             )

    def __upd_fractions(self) -> None:
        """Update fractions combobox"""
        self.__fract_cb['values'] = list(i for i in Fractions().schemes.keys())

    def __add_fractions(self) -> None:
        """
        Add fractions to the fractions.txt, update dictionary
        """
        Fractions().schemes = (self.__fract_name.get(), self.__fract_sch.get())


class Fractions:
    __instance = None

    def __new__(cls, *args, **kwargs):
        if cls.__instance is None:
            cls.__instance = super().__new__(cls)
        return cls.__instance

    @property
    def schemes(self):
        schemes = {}
        with open('fractions.txt', 'r') as f:
            for line in f:
                key, value = line.split(': ')
                value = np.array(value.split(), dtype=float)
                schemes[key] = value
            return schemes

    @schemes.setter
    def schemes(self, scheme):
        with open('fractions.txt', 'a') as f:
            f.write(f'\n{scheme[0]}: {scheme[1]}')


if __name__ == '__main__':
    # Create Tk main window
    app = App()
    app.tk_window.mainloop()
