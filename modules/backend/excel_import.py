from tkinter import filedialog, messagebox
import openpyxl
from contextlib import contextmanager
import numpy as np
from modules.backend.storage import SampleData
from modules.backend.weights_indices_calc import WeightsAndIndices


class ImportExcel:
    __db = None

    def __init__(self, db, cfg):
        self.__cfg = cfg
        self.__db = db
        self.fractions = None
        self.title = None

    @contextmanager
    def __open_wb(self):
        wb = openpyxl.load_workbook(filedialog.askopenfilename(title='Open Excel book',
                                                               filetypes=[("Excel files", ".xlsx .xls")]),
                                    data_only=True)
        try:
            yield wb
        finally:
            wb.close()

    def import_worksheet(self):
        with self.__open_wb() as wb:
            # Open active sheet
            worksheet = wb.active
            self.title = worksheet.title[:]
            # Read fractions
            self.fractions = np.array([col[0].value for col in worksheet.iter_cols(9, worksheet.max_column)],
                                      dtype=float)
            # Read rows
            for i in range(1, worksheet.max_row):
                # Read each column in row, write to row
                row = [col[i].value for col in worksheet.iter_cols(1, worksheet.max_column)]
                if row[2] in self.__db.get_data('sample'):
                    ask = messagebox.askyesnocancel(title='Conflict',
                                                    message=f'Sample{row[2]} already exists in database. '
                                                            f'Do you want to rename it to {row[2]}-{worksheet.title}?')
                    if ask:
                        row[2] = f'{row[2]}-{worksheet.title}'
                        if row[2] in self.__db.get_data('sample'):
                            messagebox.showerror(title='Conflict',
                                                 message=f'Sample{row[2]} already exists in database.')
                    elif ask is None:
                        break
                    else:
                        continue
                # Write weights
                weights = np.array(row[8:worksheet.max_column], dtype=float)
                # Write sample information to SampleData dataclass
                info = SampleData(*row[0:7], row[7].strftime("%Y-%m-%d"))
                self.__db.add_to_db(*WeightsAndIndices(self.fractions, weights, self.__cfg).get(), info)
